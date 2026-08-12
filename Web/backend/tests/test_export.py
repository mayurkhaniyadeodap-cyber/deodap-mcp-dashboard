"""Export date-range regression test: the bills export must span the WHOLE range
(all pages / all dates), not just the oldest page (one date)."""

import asyncio
import re
from types import SimpleNamespace

import pytest

from app.services import export_service


def _orders(start_id: int, day: str) -> list[dict]:
    return [
        {
            "id": start_id + i, "awb": f"A{start_id + i}", "order_date": day,
            "shipping_company": "BlueDart", "courier_slug": "blue_dart",
            "total_weight_kg": 1.0, "actual_weight_kg": 0.9, "applied_courier_rate": 50.0,
            "cod_total": 0, "customer_state": "Maharashtra", "status": "Delivered",
        }
        for i in range(500)
    ]


def test_bills_export_covers_full_range_not_one_day(monkeypatch):
    # list_orders is date-desc: offset 0 = newest 500 (2026-07-30), offset 500 =
    # oldest 500 (2026-07-01). total_matched=1000 → 2 pages. The export must include
    # BOTH → records from multiple dates spanning the selected range.
    page_new = {"total_matched": 1000, "orders": _orders(1, "2026-07-30")}
    page_old = {"total_matched": 1000, "orders": _orders(501, "2026-07-01")}

    async def _call(name, arguments=None):
        assert name == "list_orders"
        offset = (arguments or {}).get("offset", 0)
        payload = page_new if offset == 0 else page_old
        return SimpleNamespace(content=[], structuredContent=payload)

    monkeypatch.setattr("app.services.mcp_client.call_tool", _call)

    content, media, filename, count = asyncio.run(
        export_service.render("bills", "csv", "2026-07-01", "2026-07-30")
    )
    text = content.decode("utf-8-sig")

    assert "2026-07-01" in text  # oldest date (start of range) — was the ONLY date before the fix
    assert "2026-07-30" in text  # newest date (end of range) — proves the full span
    assert text.strip().count("\n") >= 1000  # ~1000 data rows → not capped to one 500-row page
    assert count >= 1000
    assert filename == "deodap_bills_2026-07-01_2026-07-30.csv"  # base range name (stamp added at HTTP layer)


def test_bills_export_xlsx_same_range(monkeypatch):
    page_new = {"total_matched": 1000, "orders": _orders(1, "2026-07-30")}
    page_old = {"total_matched": 1000, "orders": _orders(501, "2026-07-01")}

    async def _call(name, arguments=None):
        offset = (arguments or {}).get("offset", 0)
        return SimpleNamespace(content=[], structuredContent=(page_new if offset == 0 else page_old))

    monkeypatch.setattr("app.services.mcp_client.call_tool", _call)
    content, media, filename, count = asyncio.run(
        export_service.render("bills", "xlsx", "2026-07-01", "2026-07-30")
    )
    assert media == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert content[:2] == b"PK"  # a real xlsx (zip) file
    assert filename.endswith("_2026-07-01_2026-07-30.xlsx")


# --- HTTP-level: the downloaded filename carries the SELECTED range + download stamp ---
_STAMP = r"_downloaded_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}"


def test_export_csv_filename_range_and_stamp_30d(client, mock_mcp):
    mock_mcp()
    r = client.get("/api/export/csv", params={"dataset": "bills", "from": "2026-07-01", "to": "2026-07-30"})
    assert r.status_code == 200
    cd = r.headers["content-disposition"]
    assert re.search(rf'filename="deodap_bills_2026-07-01_2026-07-30{_STAMP}\.csv"', cd), cd


def test_export_xlsx_filename_range_and_stamp_7d(client, mock_mcp):
    # A different (7-day) range → the filename must reflect THOSE dates, not a fixed 30.
    mock_mcp()
    r = client.get("/api/export/xlsx", params={"dataset": "couriers", "from": "2026-07-24", "to": "2026-07-30"})
    assert r.status_code == 200
    cd = r.headers["content-disposition"]
    assert re.search(rf'filename="deodap_couriers_2026-07-24_2026-07-30{_STAMP}\.xlsx"', cd), cd


def test_export_custom_range_filename(client, mock_mcp):
    # Custom range → exact custom from/to in the filename.
    mock_mcp()
    r = client.get("/api/export/csv", params={"dataset": "zones", "from": "2026-05-03", "to": "2026-06-17"})
    assert r.status_code == 200
    assert re.search(rf'deodap_zones_2026-05-03_2026-06-17{_STAMP}\.csv', r.headers["content-disposition"])


# --- Export History -----------------------------------------------------------
def test_export_saved_to_history_and_redownload_no_mcp(client, mock_mcp, monkeypatch):
    mock_mcp()
    r = client.get("/api/export/csv", params={"dataset": "bills", "from": "2026-07-01", "to": "2026-07-30"})
    assert r.status_code == 200
    original = r.content

    items = client.get("/api/exports").json()["items"]
    assert len(items) == 1
    e = items[0]
    assert e["status"] == "completed" and e["dataset"] == "bills"
    assert e["date_from"] == "2026-07-01" and e["date_to"] == "2026-07-30"
    assert e["fmt"] == "csv" and e["record_count"] >= 1
    assert "_downloaded_" in e["filename"] and e["size_bytes"] > 0

    # Re-download must NOT call MCP — any tool call now raises.
    async def _boom(*a, **k):
        raise AssertionError("history re-download must not call MCP")

    monkeypatch.setattr("app.services.mcp_client.call_tool", _boom)
    dl = client.get(f"/api/exports/{e['id']}/download")
    assert dl.status_code == 200
    assert dl.content == original  # byte-identical to the first download
    assert e["filename"] in dl.headers["content-disposition"]


def test_history_persists_for_fresh_client_after_refresh(client, mock_mcp):
    """A brand-new client (simulating a browser refresh — fresh JS/query cache) must
    still see the persisted history from GET /api/exports (DB is the source of truth)."""
    from fastapi.testclient import TestClient

    from app.main import app

    mock_mcp()
    client.get("/api/export/csv", params={"dataset": "bills", "from": "2026-07-01", "to": "2026-07-30"})

    fresh = TestClient(app)  # reuses the same active get_current_user override
    items = fresh.get("/api/exports").json()["items"]
    assert len(items) == 1 and items[0]["dataset"] == "bills" and items[0]["status"] == "completed"


def test_history_download_file_unavailable(client, mock_mcp):
    from app.services import export_history_service

    mock_mcp()
    client.get("/api/export/csv", params={"dataset": "bills", "from": "2026-07-01", "to": "2026-07-30"})
    e = client.get("/api/exports").json()["items"][0]
    (export_history_service.EXPORT_DIR / e["filename"]).unlink()  # file gone
    dl = client.get(f"/api/exports/{e['id']}/download")
    assert dl.status_code == 404  # "File unavailable" — never falls back to MCP


def test_all_data_xlsx_has_actual_rows(client, mock_mcp):
    import io as _io

    from openpyxl import load_workbook

    mock_mcp()
    r = client.get("/api/export/xlsx", params={"dataset": "all", "from": "2026-07-01", "to": "2026-07-30"})
    assert r.status_code == 200
    wb = load_workbook(_io.BytesIO(r.content))
    # The bug: every sheet had only the header row (max_row == 1). Now sheets have data.
    populated = [ws.title for ws in wb.worksheets if ws.max_row > 1]
    assert populated, "All Data XLSX still contains only headers"
    assert "Bills" in populated  # per-order rows across the range are present

    e = client.get("/api/exports").json()["items"][0]
    assert e["dataset"] == "all" and e["status"] == "completed" and e["record_count"] > 0
    assert e["date_from"] == "2026-07-01" and e["date_to"] == "2026-07-30"


def test_all_data_csv_has_actual_rows(client, mock_mcp):
    mock_mcp()
    r = client.get("/api/export/csv", params={"dataset": "all", "from": "2026-07-01", "to": "2026-07-30"})
    assert r.status_code == 200
    lines = r.content.decode("utf-8-sig").splitlines()
    assert len(lines) > 1  # header + at least one data row


def test_all_data_history_redownload_no_mcp(client, mock_mcp, monkeypatch):
    mock_mcp()
    r = client.get("/api/export/xlsx", params={"dataset": "all", "from": "2026-07-01", "to": "2026-07-30"})
    original = r.content
    e = client.get("/api/exports").json()["items"][0]

    async def _boom(*a, **k):
        raise AssertionError("All Data history re-download must not call MCP")

    monkeypatch.setattr("app.services.mcp_client.call_tool", _boom)
    dl = client.get(f"/api/exports/{e['id']}/download")
    assert dl.status_code == 200
    assert dl.content == original  # byte-identical stored file


def test_failed_export_recorded_as_failed(client, monkeypatch):
    async def _boom(*a, **k):
        raise RuntimeError("render blew up (test)")

    monkeypatch.setattr("app.services.export_service.render", _boom)
    # The endpoint records the failure THEN re-raises (Starlette → 500 in prod; the
    # TestClient surfaces the raised error). The 'failed' row must already be saved.
    with pytest.raises(RuntimeError):
        client.get("/api/export/csv", params={"dataset": "bills", "from": "2026-07-01", "to": "2026-07-30"})

    items = client.get("/api/exports").json()["items"]
    assert len(items) == 1
    assert items[0]["status"] == "failed" and items[0]["error"]


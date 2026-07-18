"""Export service — builds CSV/XLSX byte streams for the selected date range.

Phase 2.1: the row-source is no longer static mock JSON. Each dataset is routed
through the SAME service the dashboard uses (bills_service, cod_service, …) with
the current from/to, so exporting Today vs Last-7-days vs Last-30-days yields
genuinely different files. The CSV/XLSX rendering is unchanged.

Bills note: list_orders is per-order, date-desc only, and huge (~155k rows for
30 days at ~8s / 500 rows → pulling the whole range is infeasible for a download
button). So the bills export is anchored at the START of the selected range (its
oldest page), which is fast (1-2 MCP calls) and makes every range's file
distinct. total_matched is logged so the full range size stays visible.
"""

import asyncio
import csv
import io
import logging
import time

from openpyxl import Workbook

from app.schemas.export import ExportCatalog, ExportDataset
from app.services import (
    bills_service,
    cod_service,
    courier_service,
    discrepancy_service,
    live_support,
    zone_service,
)
from app.utils.mock import load_mock

logger = logging.getLogger("live")

# One list_orders page (~8s live). Keeps the bills export responsive; the export
# is anchored at the range start so different ranges stay distinct even capped.
_EXPORT_MAX_ROWS = 500

# Rendered-file cache. Key MUST include dataset+fmt+from+to so a different range
# never returns a stale identical file. {key: (monotonic_ts, (bytes, media, name))}
_render_cache: dict = live_support.new_cache()

# dataset key -> (mock file for catalog counts, extractor, headers, human label + description)
_DATASETS: dict[str, dict] = {
    "bills": {
        "file": "bills.json",
        "extract": lambda d: d,
        "headers": ["id", "awb", "courier", "date", "weight", "zone", "amount", "cod", "status"],
        "label": "Bills",
        "description": "All courier bills with AWB, weight, zone, amount and status.",
    },
    "couriers": {
        "file": "couriers.json",
        "extract": lambda d: d,
        "headers": ["name", "code", "shipments", "freight", "rto", "cod_value", "avg_cost", "rto_pct", "on_time_pct"],
        "label": "Courier Comparison",
        "description": "Per-courier shipments, cost (freight + RTO), COD value and rates.",
    },
    "cod": {
        "file": "cod.json",
        "extract": lambda d: d["reconciliation"],
        "headers": ["courier", "orders", "cod_value"],
        "label": "COD by Courier",
        "description": "COD value and order count per courier.",
    },
    "discrepancies": {
        "file": "discrepancies.json",
        "extract": lambda d: d["rto"],
        "headers": ["courier", "orders", "count", "rate_pct"],
        "label": "RTO by Courier",
        "description": "Per-courier RTO count and rate.",
    },
    "zones": {
        "file": "zones.json",
        "extract": lambda d: d["states"],
        "headers": ["state", "orders", "total_cost", "avg_cost", "fwd_cost", "rto_cost", "delivery_rate_pct", "rto_rate_pct", "ndr_rate_pct", "avg_delivery_days"],
        "label": "State Analysis",
        "description": "Per-state shipping cost and delivery performance.",
    },
}

FORMATS = ["csv", "xlsx"]


def catalog() -> ExportCatalog:
    """Available datasets (+ nominal row counts) and formats for the Export page."""
    datasets = []
    total = 0
    for key, cfg in _DATASETS.items():
        rows = cfg["extract"](load_mock(cfg["file"]))
        total += len(rows)
        datasets.append(
            ExportDataset(key=key, label=cfg["label"], description=cfg["description"], rows=len(rows))
        )
    # Composite: every dataset in one file (CSV combined w/ a Dataset column; XLSX
    # one sheet each). Nominal count = sum of the parts.
    datasets.append(ExportDataset(
        key="all", label="All Data",
        description="Complete dashboard export containing all available live MCP data.",
        rows=total,
    ))
    return ExportCatalog(datasets=datasets, formats=FORMATS)


def is_valid_dataset(dataset: str) -> bool:
    return dataset == "all" or dataset in _DATASETS


async def _bills_rows(date_from: str | None, date_to: str | None) -> list[dict]:
    """Bills for the range, anchored at its oldest page (see module docstring)."""
    first = await bills_service.list_bills(
        page=1, page_size=_EXPORT_MAX_ROWS, date_from=date_from, date_to=date_to
    )
    if first.total_pages <= 1:
        page = first
    else:
        # Oldest page = start of the selected range → distinct per range.
        page = await bills_service.list_bills(
            page=first.total_pages, page_size=_EXPORT_MAX_ROWS,
            date_from=date_from, date_to=date_to,
        )
    # list_orders is date-desc within a page; reverse to chronological ascending.
    rows = [b.model_dump(mode="json") for b in reversed(page.items)]
    logger.info(
        "export bills: from=%s to=%s total_matched=%s exported_rows=%s span=%s..%s",
        date_from, date_to, first.total, len(rows),
        rows[0]["date"] if rows else "-", rows[-1]["date"] if rows else "-",
    )
    return rows


async def _dataset_rows(dataset: str, date_from: str | None, date_to: str | None) -> list[dict]:
    """Route each dataset through its real service with the date range."""
    if dataset == "bills":
        return await _bills_rows(date_from, date_to)

    if dataset == "couriers":
        couriers = await courier_service.list_couriers(date_from=date_from, date_to=date_to)
        rows = [c.model_dump(mode="json") for c in couriers]
    elif dataset == "cod":
        resp = await cod_service.get_cod(date_from=date_from, date_to=date_to)
        rows = [r.model_dump(mode="json") for r in resp.reconciliation]
    elif dataset == "discrepancies":
        resp = await discrepancy_service.get_discrepancies(date_from=date_from, date_to=date_to)
        rows = [d.model_dump(mode="json") for d in resp.rto]
    elif dataset == "zones":
        resp = await zone_service.get_zones(date_from=date_from, date_to=date_to)
        rows = [z.model_dump(mode="json") for z in resp.states]
    else:  # unreachable — guarded by is_valid_dataset upstream
        rows = []

    logger.info(
        "export %s: from=%s to=%s rows=%s", dataset, date_from, date_to, len(rows)
    )
    return rows


# "All Data" sheet layout (sheet name, source, headers). Sources are the SAME
# services the individual datasets use — no duplicated fetch logic.
_WEIGHT_HEADERS = ["courier", "awb", "expected_weight_kg", "billed_weight_kg", "weight_diff_kg", "status"]
_RECONCILED_HEADERS = ["courier", "reconciled_lines", "reconciled_amount"]


async def _all_sections(date_from: str | None, date_to: str | None) -> list[tuple[str, list[str], list[dict]]]:
    """Fetch every section concurrently through its real service (live only)."""
    bills, couriers, cod, disc, zones, recon = await asyncio.gather(
        _dataset_rows("bills", date_from, date_to),
        _dataset_rows("couriers", date_from, date_to),
        _dataset_rows("cod", date_from, date_to),
        _dataset_rows("discrepancies", date_from, date_to),
        _dataset_rows("zones", date_from, date_to),
        discrepancy_service.get_reconciliation(date_from=date_from, date_to=date_to),
    )
    return [
        ("Bills", _DATASETS["bills"]["headers"], bills),
        ("Courier Comparison", _DATASETS["couriers"]["headers"], couriers),
        ("COD Reconciliation", _DATASETS["cod"]["headers"], cod),
        ("RTO Analysis", _DATASETS["discrepancies"]["headers"], disc),
        ("Weight Discrepancies", _WEIGHT_HEADERS, [w.model_dump(mode="json") for w in recon.weight_disputes]),
        ("Reconciliation Summary", _RECONCILED_HEADERS, [r.model_dump(mode="json") for r in recon.reconciled]),
        ("State Analysis", _DATASETS["zones"]["headers"], zones),
    ]


def _render_all_bytes(sections: list[tuple[str, list[str], list[dict]]], fmt: str) -> tuple[bytes, str]:
    if fmt == "csv":
        # One flat file: a leading "Dataset" column identifies the section; the
        # rest is the union of every section's headers (blank where N/A).
        union: list[str] = []
        for _, headers, _ in sections:
            for h in headers:
                if h not in union:
                    union.append(h)
        fieldnames = ["Dataset", *union]
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for sheet, headers, rows in sections:
            for row in rows:
                writer.writerow({"Dataset": sheet, **{h: row.get(h) for h in headers}})
        return buf.getvalue().encode("utf-8-sig"), "text/csv"

    # XLSX: one sheet per section.
    wb = Workbook()
    wb.remove(wb.active)
    for sheet, headers, rows in sections:
        ws = wb.create_sheet(title=sheet[:31])
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _filename(dataset: str, fmt: str, date_from: str | None, date_to: str | None) -> str:
    if date_from and date_to:
        return f"deodap_{dataset}_{date_from}_{date_to}.{fmt}"
    return f"deodap_{dataset}.{fmt}"


def _render_bytes(headers: list[str], rows: list[dict], fmt: str, sheet: str) -> tuple[bytes, str]:
    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        return buf.getvalue().encode("utf-8-sig"), "text/csv"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    out = io.BytesIO()
    wb.save(out)
    return (
        out.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


async def render(
    dataset: str,
    fmt: str,
    date_from: str | None = None,
    date_to: str | None = None,
) -> tuple[bytes, str, str]:
    """Return (content, media_type, filename) for the dataset+range in the given format."""
    key = (dataset, fmt, date_from, date_to)
    now = time.monotonic()
    cached = _render_cache.get(key)
    if cached is not None and (now - cached[0]) < live_support.CACHE_TTL_SECONDS:
        return cached[1]

    if dataset == "all":
        sections = await _all_sections(date_from, date_to)
        content, media_type = _render_all_bytes(sections, fmt)
        filename = _filename("all_data", fmt, date_from, date_to)
    else:
        headers = _DATASETS[dataset]["headers"]
        rows = await _dataset_rows(dataset, date_from, date_to)
        filename = _filename(dataset, fmt, date_from, date_to)
        content, media_type = _render_bytes(headers, rows, fmt, dataset)

    result = (content, media_type, filename)
    _render_cache[key] = (now, result)
    return result
